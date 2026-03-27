"""Scoring engine — transforms LasoData into scored dimensions with alerts."""

import logging
from pathlib import Path

from openpyxl import load_workbook
from slugify import slugify

from app.constants import DIMENSIONS, HOUSE_WEIGHTS, SUMMARY_AGE_WEIGHTS, DIMENSION_LABELS
from app.models.schemas import (
    StarRow,
    ScorePoint,
    Alert,
    DimensionScores,
    ScoringResult,
)

logger = logging.getLogger(__name__)

class ScoringError(Exception):
    """Raised when scoring input data is invalid."""
    pass


# Dimension columns that have alert tag columns in the xlsx.
# van_menh has NO alerts so it has no tag columns.
_ALERT_DIMENSIONS = [d for d in DIMENSIONS if d != "van_menh"]


class ScoringEngine:
    """Stateless scoring calculator. Load xlsx once, reuse across requests."""

    def __init__(self, xlsx_path: str = "data/laso_points.xlsx"):
        self._star_table: dict[str, StarRow] = self._load_star_table(xlsx_path)

    def _load_star_table(self, path: str) -> dict[str, StarRow]:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["laso_points"]

        # Read header row to build column index
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h else "" for h in next(rows)]
        col = {name: idx for idx, name in enumerate(header) if name}

        def _safe_get(row: tuple, idx: int):
            """Safely get a value from a row tuple, returning None if out of range."""
            return row[idx] if idx < len(row) else None

        table: dict[str, StarRow] = {}

        for row_vals in rows:
            sao_raw = _safe_get(row_vals, col["sao"])
            if sao_raw is None:
                continue

            slug = slugify(str(sao_raw).lower())
            if slug in table:
                logger.warning("Duplicate star slug '%s' (name: %s) — keeping first", slug, sao_raw)
                continue

            point_val = _safe_get(row_vals, col["point"])
            point = int(point_val) if point_val is not None else 0

            # Build dimension weights — empty/NaN -> 1.0
            weights: dict[str, float] = {}
            for dim in DIMENSIONS:
                val = _safe_get(row_vals, col[dim]) if dim in col else None
                if val is None or val == "":
                    weights[dim] = 1.0
                else:
                    try:
                        weights[dim] = float(val)
                    except (ValueError, TypeError):
                        weights[dim] = 1.0

            # Build pct_fvg
            pct_fvg: dict[int, str | None] = {}
            for level in (30, 50):
                fvg_col = f"pct_fvg_{level}"
                if fvg_col in col:
                    val = _safe_get(row_vals, col[fvg_col])
                    if val and str(val).strip() in ("pos", "neg"):
                        pct_fvg[level] = str(val).strip()
                    else:
                        pct_fvg[level] = None
                else:
                    pct_fvg[level] = None

            # Build alert tags per dimension per level
            alert_tags: dict[str, dict[int, str | None]] = {}
            for dim in _ALERT_DIMENSIONS:
                dim_tags: dict[int, str | None] = {}
                for level in (30, 50):
                    tag_col = f"{dim}_tag_{level}"
                    if tag_col in col:
                        val = _safe_get(row_vals, col[tag_col])
                        if val and str(val).strip():
                            dim_tags[level] = str(val).strip()
                        else:
                            dim_tags[level] = None
                    else:
                        dim_tags[level] = None
                alert_tags[dim] = dim_tags

            table[slug] = StarRow(
                slug=slug,
                name=str(sao_raw),
                point=point,
                weights=weights,
                alert_tags=alert_tags,
                pct_fvg=pct_fvg,
            )

        wb.close()
        logger.info("Loaded %d stars from %s", len(table), path)
        return table

    def _build_raw_scores(
        self, cungs: list, dim: str
    ) -> dict[str, tuple[float, float]]:
        """Calculate raw pos/neg scores per cung for a dimension.

        Args:
            cungs: List of Cung (or any object with .name and .stars[].slug).
            dim: Dimension key (e.g., "su_nghiep").

        Returns:
            Dict mapping lowered cung name -> (raw_pos, raw_neg).
        """
        result: dict[str, tuple[float, float]] = {}

        for cung in cungs:
            raw_pos = 0.0
            raw_neg = 0.0

            for star in cung.stars:
                row = self._star_table.get(star.slug)
                if row is None:
                    logger.warning(
                        "Star not found in table: slug='%s' (cung='%s')",
                        star.slug,
                        cung.name,
                    )
                    continue

                weight = row.weights.get(dim, 1.0)
                contribution = row.point * weight

                if row.point > 0:
                    raw_pos += contribution
                else:
                    raw_neg += contribution

            result[cung.name.lower()] = (raw_pos, raw_neg)

        return result

    def _calc_anchor(
        self,
        lifetime_raws: dict[str, tuple[float, float]],
        dim: str,
        than_cu: str,
    ) -> tuple[float, float]:
        """Calculate house weighting anchor from lifetime raw scores.

        Args:
            lifetime_raws: Dict from _build_raw_scores (cung_name_lower -> (pos, neg)).
            dim: Dimension key.
            than_cu: Which cung "Than" resides in (e.g., "Quan Loc").

        Returns:
            (anchor_pos, anchor_neg) tuple.
        """
        anchor_pos = 0.0
        anchor_neg = 0.0

        for cung_name, house_weight in HOUSE_WEIGHTS[dim]:
            resolved = than_cu.lower() if cung_name == "thân" else cung_name
            raw_pos, raw_neg = lifetime_raws.get(resolved, (0.0, 0.0))
            anchor_pos += raw_pos * house_weight
            anchor_neg += raw_neg * house_weight

        return anchor_pos, anchor_neg

    def _calc_final(
        self,
        raws: dict[str, tuple[float, float]],
        anchor: tuple[float, float],
        periods: list[str],
        cung_order: list[str],
    ) -> list[ScorePoint]:
        """Calculate final scores for a list of cungs.

        Args:
            raws: cung_name_lower -> (raw_pos, raw_neg).
            anchor: (anchor_pos, anchor_neg) from _calc_anchor.
            periods: X-axis labels in order.
            cung_order: Cung names (lowered) matching periods order.

        Returns:
            List of ScorePoint, one per period.
        """
        anchor_pos, anchor_neg = anchor
        points: list[ScorePoint] = []

        for period, cung_name in zip(periods, cung_order):
            raw_pos, raw_neg = raws.get(cung_name, (0.0, 0.0))
            final_pos = (raw_pos + anchor_pos) / 2
            final_neg = (raw_neg + anchor_neg) / 2
            final_tb = final_pos + final_neg

            points.append(ScorePoint(
                period=str(period),
                duong=round(final_pos, 2),
                am=round(final_neg, 2),
                tb=round(final_tb, 2),
            ))

        return points

    def _summary_score(self, lifetime_points: list[ScorePoint]) -> float:
        """Weighted average of lifetime tb values using SUMMARY_AGE_WEIGHTS."""
        ages = list(range(0, 120, 10))
        if len(lifetime_points) != len(ages):
            logger.error(
                "Expected %d lifetime points, got %d", len(ages), len(lifetime_points)
            )
            return 0.0

        weighted_sum = 0.0
        weight_total = 0.0

        for point, age in zip(lifetime_points, ages):
            w = SUMMARY_AGE_WEIGHTS.get(age, 1.0)
            weighted_sum += point.tb * w
            weight_total += w

        return round(weighted_sum / weight_total, 2) if weight_total > 0 else 0.0

    def _detect_alerts(
        self,
        points: list[ScorePoint],
        cungs: list,
        dim: str,
        timeframe: str,
    ) -> list[Alert]:
        """Detect score change alerts between consecutive periods.

        Args:
            points: ScorePoint list (ordered by period).
            cungs: Matching cung list (same order as points) -- stars used for tag lookup.
            dim: Dimension key.
            timeframe: "lifetime", "decade", or "monthly".

        Returns:
            List of Alert objects.
        """
        alerts: list[Alert] = []

        # Determine age cutoff for lifetime charts
        max_age = None
        if timeframe == "lifetime":
            max_age = 60 if dim == "con_cai" else 80

        for col_attr, col_name in [("duong", "pos"), ("am", "neg")]:
            values = [getattr(p, col_attr) for p in points]

            for i in range(1, len(values)):
                prev = values[i - 1]
                curr = values[i]

                # Skip if previous is zero
                if abs(prev) < 1e-9:
                    continue

                pct_change = ((curr - prev) / abs(prev)) * 100

                # Skip after age cutoff on lifetime
                if max_age is not None and timeframe == "lifetime":
                    try:
                        age = int(points[i].period.split("-")[0])
                        if age >= max_age:
                            continue
                    except (ValueError, IndexError):
                        pass

                # Determine alert level
                abs_pct = abs(pct_change)
                if abs_pct >= 50:
                    level = 50
                elif abs_pct >= 30:
                    level = 30
                else:
                    continue

                direction = "pos" if pct_change > 0 else "neg"

                # Find matching stars in destination cung
                if i < len(cungs):
                    for star in cungs[i].stars:
                        row = self._star_table.get(star.slug)
                        if row is None:
                            continue

                        # Check pct_fvg direction match
                        if row.pct_fvg.get(level) != direction:
                            continue

                        # Get tag text
                        tag = row.alert_tags.get(dim, {}).get(level)
                        if not tag:
                            continue

                        alerts.append(Alert(
                            type="positive" if direction == "pos" else "negative",
                            dimension=dim,
                            period=points[i].period,
                            tag=tag,
                            level=level,
                            star_name=row.slug,
                        ))

        return alerts

    def score(self, laso, nam_xem: int = 2026) -> ScoringResult:
        """Main entry point: LasoData -> ScoringResult.

        Args:
            laso: LasoData (or compatible object with cung, cung_10yrs, cung_12months, than_cu).
            nam_xem: Year being analyzed (for decade/monthly x-axis labels).

        Returns:
            ScoringResult with 8 dimensions scored.
        """
        # Validate input
        if len(laso.cung) != 12:
            raise ScoringError(f"Expected 12 lifetime cungs, got {len(laso.cung)}")
        if len(laso.cung_12months) < 12:
            raise ScoringError(f"Expected 12 monthly cungs, got {len(laso.cung_12months)}")

        than_cu = laso.than_cu

        dimensions: dict[str, DimensionScores] = {}
        all_alerts: list[Alert] = []

        for dim in DIMENSIONS:
            # Step 1: Build raw scores for all 3 timeframes
            lifetime_raws = self._build_raw_scores(laso.cung, dim)
            decade_raws = self._build_raw_scores(laso.cung_10yrs, dim)
            monthly_raws = self._build_raw_scores(laso.cung_12months, dim)

            # Step 2: Anchor from LIFETIME data only
            anchor = self._calc_anchor(lifetime_raws, dim, than_cu)

            # Step 3: Final scores -- Lifetime (12 points)
            lifetime_cung_order = [c.name.lower() for c in laso.cung]
            lifetime_periods = [f"{a}-{a+10}" for a in range(0, 120, 10)]
            lifetime_points = self._calc_final(
                lifetime_raws, anchor, lifetime_periods, lifetime_cung_order
            )

            # Step 3: Final scores -- Decade (first 10 of 12 cungs)
            decade_cungs = laso.cung_10yrs[:-2]
            decade_cung_order = [c.name.lower() for c in decade_cungs]
            decade_periods = [str(y) for y in range(nam_xem, nam_xem + 10)]
            decade_points = self._calc_final(
                decade_raws, anchor, decade_periods, decade_cung_order
            )

            # Step 3: Final scores -- Monthly (13 points, first prepended)
            monthly_cung_order = [laso.cung_12months[0].name.lower()]
            monthly_cung_order += [mc.name.lower() for mc in laso.cung_12months]
            monthly_periods = [f"Th.{laso.cung_12months[0].month}/{nam_xem}"]
            monthly_periods += [
                f"Th.{mc.month}/{nam_xem}" for mc in laso.cung_12months
            ]
            monthly_points = self._calc_final(
                monthly_raws, anchor, monthly_periods, monthly_cung_order
            )

            # Step 4: Alerts (skip for van_menh)
            alerts: list[Alert] = []
            if dim != "van_menh":
                # Lifetime alerts
                lifetime_alert_cungs = list(laso.cung)
                alerts.extend(
                    self._detect_alerts(lifetime_points, lifetime_alert_cungs, dim, "lifetime")
                )
                # Decade alerts
                alerts.extend(
                    self._detect_alerts(decade_points, list(decade_cungs), dim, "decade")
                )
                # Monthly alerts
                monthly_alert_cungs = [laso.cung_12months[0]] + list(laso.cung_12months)
                alerts.extend(
                    self._detect_alerts(monthly_points, monthly_alert_cungs, dim, "monthly")
                )

            # Step 5: Summary score
            summary = self._summary_score(lifetime_points)

            dimensions[dim] = DimensionScores(
                dimension=dim,
                label=DIMENSION_LABELS[dim],
                lifetime=lifetime_points,
                decade=decade_points,
                monthly=monthly_points,
                alerts=alerts,
                summary_score=summary,
            )
            all_alerts.extend(alerts)

        # Sort all_alerts: level desc, then by dimension order
        dim_order = {d: i for i, d in enumerate(DIMENSIONS)}
        all_alerts.sort(key=lambda a: (-a.level, dim_order.get(a.dimension, 99)))

        return ScoringResult(dimensions=dimensions, all_alerts=all_alerts)
